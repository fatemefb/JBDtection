"""
siemens_mode.py — Siemens platform mode for Control IntelliCraft.

REUSES DataAnalysisModule's TagJBExtractor for:
  • build_tag_vectors_from_excel() — IO List loading + vector matcher building
  • vector_matcher.find_similar_tags() — fuzzy tag matching
  • _normalize_ocr_tag_candidate() — text normalization (handles unicode dashes, etc.)

WHAT IT DOES:
  1. Loads IO List tags via the existing TagJBExtractor
  2. For each PDF page, extracts digital text via page.get_text("words")
  3. AGGRESSIVE multi-strategy tag matching:
     - Phase 1: Direct exact match (case-insensitive)
     - Phase 2: Normalized match (strip non-alphanumeric, handle unicode dashes)
     - Phase 3: Multi-word combination (2-3 adjacent words concatenated)
     - Phase 4: Substring match (IO tag found as substring of a PDF word)
     - Phase 5: Fuzzy match via vector_matcher.find_similar_tags() (threshold ≥ 0.85)
  4. Draws colored vector bounding boxes on the PDF page (text stays selectable)
  5. Saves annotated PDF + multi-sheet Excel report + ZIP bundle

WHAT IT DOES NOT DO (per spec):
  ✗ No JB/MC/Spare/Cable detection (empty patterns set)
  ✗ No OCR (digital text layer only)
  ✗ No PDFClassifier
  ✗ No page filtering based on JB presence
  ✗ No tag deletion based on JB presence
"""

from __future__ import annotations

import logging
import os
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)


# ── Constants ──────────────────────────────────────────────────────────────
_TAG_COLUMN_CANDIDATES = (
    "tag no", "tag_no", "tag_number", "tag-number", "tagno", "tag", "تگ",
)

_BBOX_COLORS = [
    (1.0, 0.30, 0.20),   # red-orange
    (0.20, 0.55, 1.0),   # blue
    (0.20, 0.80, 0.40),  # green
    (1.0, 0.75, 0.20),   # amber
    (0.65, 0.30, 1.0),   # purple
    (0.10, 0.70, 0.85),  # cyan
]

_BBOX_STROKE_WIDTH = 1.4
_LABEL_FONT_SIZE = 7

# Fuzzy match threshold for vector_matcher.find_similar_tags()
# 0.85 = allow ~15% character difference (handles OCR noise, unicode issues)
_FUZZY_THRESHOLD = 0.85

# Max adjacent words to try concatenating (3 = try pairs and triples)
_MAX_WORD_COMBINATION = 3

# Min tag length to attempt matching (skip 1-2 char tokens)
_MIN_TAG_LENGTH = 3


# ── Result dataclass ──────────────────────────────────────────────────────
@dataclass
class SiemensMatch:
    pdf_filename: str
    page_number: int
    tag: str
    matched_text: str
    bbox: Tuple[float, float, float, float]
    match_type: str  # 'exact', 'normalized', 'multiword', 'substring', 'fuzzy'
    similarity_score: float = 1.0
    color_index: int = 0


@dataclass
class SiemensResult:
    pdf_paths: List[str]
    excel_path: str
    total_io_tags: int
    matched_tags: set = field(default_factory=set)
    unmatched_tags: set = field(default_factory=set)
    per_pdf: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    matches: List[SiemensMatch] = field(default_factory=list)
    annotated_pdfs: List[str] = field(default_factory=list)
    report_excel: Optional[str] = None
    zip_bundle: Optional[str] = None
    started_at: Optional[str] = None
    finished_at: Optional[str] = None
    duration_seconds: float = 0.0
    run_dir: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "pdf_paths": self.pdf_paths,
            "excel_path": self.excel_path,
            "total_io_tags": self.total_io_tags,
            "matched_count": len(self.matched_tags),
            "unmatched_count": len(self.unmatched_tags),
            "match_rate_pct": round(
                (len(self.matched_tags) / self.total_io_tags * 100)
                if self.total_io_tags else 0.0, 2),
            "matched_tags": sorted(self.matched_tags),
            "unmatched_tags": sorted(self.unmatched_tags),
            "per_pdf": self.per_pdf,
            "annotated_pdfs": self.annotated_pdfs,
            "report_excel": self.report_excel,
            "zip_bundle": self.zip_bundle,
            "run_dir": self.run_dir,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "duration_seconds": round(self.duration_seconds, 2),
            "total_matches": len(self.matches),
            "match_type_breakdown": self._match_type_breakdown(),
        }

    def _match_type_breakdown(self) -> Dict[str, int]:
        breakdown: Dict[str, int] = {}
        for m in self.matches:
            breakdown[m.match_type] = breakdown.get(m.match_type, 0) + 1
        return breakdown


# ── Processor ─────────────────────────────────────────────────────────────
class SiemensModeProcessor:
    """
    Wraps the existing TagJBExtractor from DataAnalysisModule.
    Reuses its IO List loading, vector matcher, and text normalization.
    Adds aggressive multi-strategy tag matching + vector bbox drawing.
    """

    def __init__(self):
        self._extractor = None
        self._io_tags_upper: Dict[str, str] = {}  # upper_tag -> original_tag
        self._io_tags_normalized: Dict[str, str] = {}  # normalized_tag -> original_tag

    def _get_extractor(self):
        """Lazily instantiate TagJBExtractor from DataAnalysisModule.
        Returns None if the module can't be imported (e.g. missing Levenshtein
        dependency in dev environments) — the 4 non-fuzzy matching strategies
        still work without it."""
        if self._extractor is not None:
            return self._extractor
        try:
            from DataAnalysisModule import TagJBExtractor
            self._extractor = TagJBExtractor()
            # Disable JB/MC/Spare/Cable detection by setting empty patterns
            self._extractor.set_patterns(
                jb_examples=[],
                mc_examples=[],
                spare_examples=[],
                cable_examples=[],
            )
            logger.info("Siemens mode: TagJBExtractor initialized with empty JB/MC patterns")
            return self._extractor
        except ImportError as exc:
            logger.warning(
                "TagJBExtractor not available (%s) — fuzzy matching disabled. "
                "Other 4 strategies (exact, normalized, multiword, substring) still work.",
                exc
            )
            self._extractor = None
            return None
        except Exception as exc:
            logger.error("Failed to initialize TagJBExtractor: %s", exc)
            self._extractor = None
            return None

    # ── Text normalization (mirrors TagJBExtractor._normalize_ocr_tag_candidate) ──

    @staticmethod
    def _normalize_tag(text: str) -> str:
        """
        Aggressive normalization: uppercase, strip whitespace, replace unicode
        dashes with ASCII dash, strip non-alphanumeric (keep dashes).
        """
        if not text:
            return ""
        # Unicode normalize (NFKD) to convert fancy unicode to ASCII equivalents
        normalized = unicodedata.normalize("NFKD", str(text))
        normalized = normalized.encode("ascii", "ignore").decode("ascii")
        normalized = normalized.strip().upper()
        # Replace various dash-like chars with standard dash
        for dash_char in ["—", "–", "‐", "‑", "‒", "−"]:
            normalized = normalized.replace(dash_char, "-")
        # Remove spaces
        normalized = re.sub(r"\s+", "", normalized)
        # Strip leading/trailing dashes and dots
        normalized = normalized.strip("-.")
        return normalized

    @staticmethod
    def _strip_alphanumeric(text: str) -> str:
        """Strip everything except A-Z, 0-9 (no dashes)."""
        if not text:
            return ""
        return re.sub(r"[^A-Z0-9]", "", str(text).upper())

    # ── Public API ────────────────────────────────────────────────────────

    def process(
        self,
        pdf_paths: List[str],
        excel_path: str,
        output_pdf_dir: str,
        output_excel_path: str,
        create_zip: bool = True,
        zip_path: Optional[str] = None,
    ) -> SiemensResult:
        started = datetime.now()
        start_ts = started.timestamp()
        logger.info("=" * 80)
        logger.info("🚀 SIEMENS MODE (v2 — DataAnalysisModule wrapper) — starting")
        logger.info("=" * 80)

        result = SiemensResult(
            pdf_paths=list(pdf_paths),
            excel_path=excel_path,
            total_io_tags=0,
            started_at=started.isoformat(),
        )

        # 1) Load IO tags using the existing TagJBExtractor (if available)
        extractor = self._get_extractor()
        if extractor is not None:
            try:
                extractor.build_tag_vectors_from_excel(excel_path)
            except Exception as exc:
                logger.warning("build_tag_vectors_from_excel failed: %s — fuzzy matching may be limited", exc)

        # Also load IO tags ourselves (for direct matching)
        io_tags = self._load_io_tags(excel_path)
        result.total_io_tags = len(io_tags)
        if not io_tags:
            logger.warning("Siemens mode: no tags loaded from %s — aborting", excel_path)
            result.finished_at = datetime.now().isoformat()
            result.duration_seconds = datetime.now().timestamp() - start_ts
            return result

        # Build lookup maps for fast matching
        for tag in io_tags:
            upper = tag.upper()
            normalized = self._normalize_tag(tag)
            alphanumeric = self._strip_alphanumeric(tag)
            self._io_tags_upper[upper] = tag
            if normalized:
                self._io_tags_normalized[normalized] = tag
            if alphanumeric:
                self._io_tags_normalized[alphanumeric] = tag

        logger.info("Loaded %d IO tags. Lookup maps: upper=%d, normalized=%d",
                     len(io_tags), len(self._io_tags_upper), len(self._io_tags_normalized))

        # 2) Prepare output directory
        os.makedirs(output_pdf_dir, exist_ok=True)

        # 3) Process each PDF
        for pdf_idx, pdf_path in enumerate(pdf_paths):
            pdf_filename = os.path.basename(pdf_path)
            logger.info("─" * 80)
            logger.info("📄 [%d/%d] %s", pdf_idx + 1, len(pdf_paths), pdf_filename)
            try:
                per_pdf_stats, annotated_pdf, page_matches = self._process_one_pdf(
                    pdf_path=pdf_path,
                    output_pdf_dir=output_pdf_dir,
                )
                result.per_pdf[pdf_filename] = per_pdf_stats
                result.annotated_pdfs.append(annotated_pdf)
                result.matches.extend(page_matches)
                for m in page_matches:
                    result.matched_tags.add(m.tag)
            except Exception as exc:
                logger.error("Failed to process %s: %s", pdf_filename, exc, exc_info=True)
                result.per_pdf[pdf_filename] = {
                    "status": "error",
                    "error": str(exc),
                    "matched_tags_count": 0,
                    "total_matches": 0,
                    "pages_processed": 0,
                }

        # 4) Compute unmatched tags
        result.unmatched_tags = set(io_tags) - result.matched_tags

        # 5) Generate match report Excel
        try:
            self._write_match_report(output_excel_path, result)
            result.report_excel = output_excel_path
            logger.info("✅ Match report saved: %s", output_excel_path)
        except Exception as exc:
            logger.error("Failed to write match report: %s", exc, exc_info=True)

        # 6) Optional ZIP bundle
        if create_zip:
            try:
                zip_file = self._create_zip_bundle(
                    output_pdf_dir=output_pdf_dir,
                    annotated_pdfs=result.annotated_pdfs,
                    report_excel=result.report_excel,
                    zip_path=zip_path,
                )
                if zip_file:
                    result.zip_bundle = zip_file
                    logger.info("📦 ZIP bundle: %s", zip_file)
            except Exception as exc:
                logger.error("Failed to create ZIP: %s", exc, exc_info=True)

        finished = datetime.now()
        result.finished_at = finished.isoformat()
        result.duration_seconds = finished.timestamp() - start_ts

        # Log summary
        breakdown = result._match_type_breakdown()
        logger.info("=" * 80)
        logger.info("✅ SIEMENS MODE — finished in %.2fs", result.duration_seconds)
        logger.info("   Total IO tags:  %d", result.total_io_tags)
        logger.info("   Matched:        %d (%.1f%%)",
                     len(result.matched_tags),
                     (len(result.matched_tags) / result.total_io_tags * 100) if result.total_io_tags else 0)
        logger.info("   Unmatched:      %d", len(result.unmatched_tags))
        logger.info("   Total matches (incl. duplicates): %d", len(result.matches))
        logger.info("   Match type breakdown: %s", breakdown)
        logger.info("=" * 80)
        return result

    # ── IO List loading ────────────────────────────────────────────────────

    def _load_io_tags(self, excel_path: str) -> List[str]:
        """Read tags from IO List Excel. Finds Tag column case-insensitively."""
        if not excel_path or not os.path.exists(excel_path):
            logger.error("IO List Excel not found: %s", excel_path)
            return []
        try:
            df = pd.read_excel(excel_path, dtype=str)
        except Exception as exc:
            logger.error("Failed to read Excel %s: %s", excel_path, exc)
            return []
        if df.empty:
            return []

        tag_col = None
        for col in df.columns:
            col_norm = str(col).strip().lower().replace(" ", "").replace("-", "_").replace(".", "")
            if col_norm in _TAG_COLUMN_CANDIDATES:
                tag_col = col
                break
        if tag_col is None:
            logger.error("Tag column not found. Looked for: %s. Available: %s",
                         _TAG_COLUMN_CANDIDATES, list(df.columns))
            return []

        tags = (
            df[tag_col]
            .dropna()
            .astype(str)
            .str.strip()
            .pipe(lambda s: s[s != ""])
            .pipe(lambda s: s[s.str.lower() != "nan"])
            .tolist()
        )
        seen = set()
        unique_tags = []
        for t in tags:
            key = t.upper()
            if key not in seen:
                seen.add(key)
                unique_tags.append(t)
        return unique_tags

    # ── Per-PDF processing ─────────────────────────────────────────────────

    def _process_one_pdf(
        self,
        pdf_path: str,
        output_pdf_dir: str,
    ) -> Tuple[Dict[str, Any], str, List[SiemensMatch]]:
        pdf_filename = os.path.basename(pdf_path)
        doc = fitz.open(pdf_path)
        if doc.is_encrypted:
            try:
                doc.authenticate("")
            except Exception:
                logger.warning("PDF %s is encrypted — skipping", pdf_filename)

        annotated_pdf_path = os.path.join(output_pdf_dir, f"annotated_{pdf_filename}")
        matches: List[SiemensMatch] = []
        pages_processed = 0
        color_idx = 0
        total_pages = len(doc)

        for page_num in range(total_pages):
            try:
                page = doc.load_page(page_num)
            except Exception as exc:
                logger.warning("Failed to load page %d of %s: %s", page_num + 1, pdf_filename, exc)
                continue

            # Get all words on this page (digital text layer)
            # Each word: (x0, y0, x1, y1, text, block_no, line_no, word_no)
            try:
                words = page.get_text("words")
            except Exception as exc:
                logger.warning("get_text('words') failed on page %d: %s", page_num + 1, exc)
                words = []

            if not words:
                logger.info("Page %d: no digital text — skipping", page_num + 1)
                continue

            pages_processed += 1

            # ── Phase 1 & 2: Single-word matching (exact + normalized) ──
            for w in words:
                x0, y0, x1, y1, text = w[0], w[1], w[2], w[3], w[4]
                text_str = str(text).strip()
                if not text_str or len(text_str) < _MIN_TAG_LENGTH:
                    continue

                matched_tag, match_type, score = self._match_single_word(text_str)
                if matched_tag is None:
                    continue

                color = _BBOX_COLORS[color_idx % len(_BBOX_COLORS)]
                color_idx += 1
                self._draw_bbox(page, x0, y0, x1, y1, matched_tag, color)
                matches.append(SiemensMatch(
                    pdf_filename=pdf_filename,
                    page_number=page_num + 1,
                    tag=matched_tag,
                    matched_text=text_str,
                    bbox=(x0, y0, x1, y1),
                    match_type=match_type,
                    similarity_score=score,
                    color_index=(color_idx - 1) % len(_BBOX_COLORS),
                ))

            # ── Phase 3: Multi-word combination matching ──
            # Try concatenating 2-3 adjacent words and matching
            matches = self._try_multiword_matches(
                page, words, pdf_filename, page_num, matches, color_idx
            )
            color_idx = len(matches)  # keep color counter in sync

            # ── Phase 4: Substring matching ──
            # For each IO tag, check if it appears as a substring of any word
            # This catches cases where the PDF has "TAG:21HS-001" and IO List has "21HS-001"
            matches = self._try_substring_matches(
                page, words, pdf_filename, page_num, matches, color_idx
            )
            color_idx = len(matches)

            # ── Phase 5: Fuzzy matching via vector_matcher ──
            # Use the existing DataAnalysisModule's vector_matcher for tags
            # that didn't match exactly but are close (OCR noise, unicode issues)
            matches = self._try_fuzzy_matches(
                page, words, pdf_filename, page_num, matches, color_idx
            )
            color_idx = len(matches)

        # Save annotated PDF
        try:
            doc.save(annotated_pdf_path, garbage=3, deflate=True)
        finally:
            doc.close()

        matched_tags_set = {m.tag for m in matches}
        per_pdf_stats = {
            "status": "ok",
            "pdf_path": pdf_path,
            "annotated_pdf": annotated_pdf_path,
            "pages_processed": pages_processed,
            "total_pages": total_pages,
            "total_matches": len(matches),
            "matched_tags_count": len(matched_tags_set),
            "matched_tags": sorted(matched_tags_set),
            "pages_with_matches": sorted({m.page_number for m in matches}),
        }
        return per_pdf_stats, annotated_pdf_path, matches

    # ── Matching strategies ────────────────────────────────────────────────

    def _match_single_word(self, text: str) -> Tuple[Optional[str], str, float]:
        """
        Try to match a single word against IO List tags.
        Returns (matched_tag, match_type, score) or (None, '', 0.0).
        """
        upper = text.upper()

        # Phase 1: Direct exact match (case-insensitive)
        if upper in self._io_tags_upper:
            return self._io_tags_upper[upper], "exact", 1.0

        # Phase 2: Normalized match
        normalized = self._normalize_tag(text)
        if normalized and normalized in self._io_tags_normalized:
            return self._io_tags_normalized[normalized], "normalized", 1.0

        # Phase 2b: Alphanumeric-only match (strip ALL non-alphanumeric)
        alphanumeric = self._strip_alphanumeric(text)
        if alphanumeric and alphanumeric != normalized and alphanumeric in self._io_tags_normalized:
            return self._io_tags_normalized[alphanumeric], "normalized", 0.98

        return None, "", 0.0

    def _try_multiword_matches(
        self,
        page,
        words: List[Tuple],
        pdf_filename: str,
        page_num: int,
        matches: List[SiemensMatch],
        color_idx: int,
    ) -> List[SiemensMatch]:
        """
        Phase 3: Try concatenating 2-3 adjacent words and matching against IO List.
        Catches tags that PyMuPDF split at dashes, spaces, or other boundaries.
        """
        if not words:
            return matches

        # Get already-matched texts on this page to avoid double-matching
        already_matched_texts = {m.matched_text for m in matches if m.page_number == page_num + 1}

        for combo_len in range(2, _MAX_WORD_COMBINATION + 1):
            for i in range(len(words) - combo_len + 1):
                combo_words = words[i:i + combo_len]

                # Check adjacency (same block, same line, consecutive word numbers)
                is_adjacent = True
                for j in range(len(combo_words) - 1):
                    w1 = combo_words[j]
                    w2 = combo_words[j + 1]
                    if w1[5] != w2[5] or w1[6] != w2[6] or w1[7] + 1 != w2[7]:
                        is_adjacent = False
                        break
                if not is_adjacent:
                    continue

                combined_text = "".join(str(w[4]) for w in combo_words)
                if combined_text in already_matched_texts:
                    continue

                # Try exact + normalized match on combined text
                matched_tag, match_type, score = self._match_single_word(combined_text)
                if matched_tag is None:
                    continue

                # Compute bounding box spanning all words in the combination
                x0 = min(w[0] for w in combo_words)
                y0 = min(w[1] for w in combo_words)
                x1 = max(w[2] for w in combo_words)
                y1 = max(w[3] for w in combo_words)

                color = _BBOX_COLORS[color_idx % len(_BBOX_COLORS)]
                color_idx += 1
                self._draw_bbox(page, x0, y0, x1, y1, matched_tag, color)
                matches.append(SiemensMatch(
                    pdf_filename=pdf_filename,
                    page_number=page_num + 1,
                    tag=matched_tag,
                    matched_text=combined_text,
                    bbox=(x0, y0, x1, y1),
                    match_type=f"multiword_{combo_len}",
                    similarity_score=score,
                    color_index=(color_idx - 1) % len(_BBOX_COLORS),
                ))
                already_matched_texts.add(combined_text)

        return matches

    def _try_substring_matches(
        self,
        page,
        words: List[Tuple],
        pdf_filename: str,
        page_num: int,
        matches: List[SiemensMatch],
        color_idx: int,
    ) -> List[SiemensMatch]:
        """
        Phase 4: For each word in the PDF, check if any IO tag appears as a substring.
        Catches cases like PDF text "TAG:21HS-001" where IO tag is "21HS-001".
        """
        if not words:
            return matches

        already_matched = {(m.matched_text, m.tag) for m in matches if m.page_number == page_num + 1}

        for w in words:
            x0, y0, x1, y1, text = w[0], w[1], w[2], w[3], w[4]
            word_text = str(text).strip()
            if not word_text or len(word_text) < _MIN_TAG_LENGTH:
                continue
            word_upper = word_text.upper()

            # Check if this word contains any IO tag as a substring
            for io_tag_upper, io_tag_orig in self._io_tags_upper.items():
                if len(io_tag_upper) < _MIN_TAG_LENGTH:
                    continue
                if io_tag_upper in word_upper and word_upper != io_tag_upper:
                    # Substring match found! But only draw if we haven't already
                    # matched this exact (word, tag) pair
                    if (word_text, io_tag_orig) in already_matched:
                        continue
                    # Check if this tag was already matched on this word via exact/normalized
                    already_exact = any(
                        m.tag == io_tag_orig and m.matched_text == word_text
                        for m in matches if m.page_number == page_num + 1
                    )
                    if already_exact:
                        continue

                    color = _BBOX_COLORS[color_idx % len(_BBOX_COLORS)]
                    color_idx += 1
                    self._draw_bbox(page, x0, y0, x1, y1, io_tag_orig, color)
                    matches.append(SiemensMatch(
                        pdf_filename=pdf_filename,
                        page_number=page_num + 1,
                        tag=io_tag_orig,
                        matched_text=word_text,
                        bbox=(x0, y0, x1, y1),
                        match_type="substring",
                        similarity_score=0.90,
                        color_index=(color_idx - 1) % len(_BBOX_COLORS),
                    ))
                    already_matched.add((word_text, io_tag_orig))

        return matches

    def _try_fuzzy_matches(
        self,
        page,
        words: List[Tuple],
        pdf_filename: str,
        page_num: int,
        matches: List[SiemensMatch],
        color_idx: int,
    ) -> List[SiemensMatch]:
        """
        Phase 5: Use the existing DataAnalysisModule's vector_matcher for fuzzy matching.
        This catches OCR noise, unicode issues, and slight character differences.
        """
        extractor = self._get_extractor()
        if extractor is None:
            return matches  # TagJBExtractor not available — skip fuzzy phase
        vector_matcher = getattr(extractor, "vector_matcher", None)
        if vector_matcher is None:
            logger.debug("vector_matcher not available — skipping fuzzy phase")
            return matches

        # Get already-matched words on this page
        already_matched_words = {m.matched_text for m in matches if m.page_number == page_num + 1}

        for w in words:
            x0, y0, x1, y1, text = w[0], w[1], w[2], w[3], w[4]
            word_text = str(text).strip()
            if not word_text or len(word_text) < _MIN_TAG_LENGTH:
                continue
            if word_text in already_matched_words:
                continue

            # Normalize the word for matching
            normalized = self._normalize_tag(word_text)
            if not normalized or len(normalized) < _MIN_TAG_LENGTH:
                continue

            try:
                similar_tags = vector_matcher.find_similar_tags(normalized)
            except Exception:
                continue

            if not similar_tags:
                continue

            best_match, best_score = similar_tags[0]
            if best_score < _FUZZY_THRESHOLD:
                continue

            # Check if this tag was already matched
            already_matched_tag = any(
                m.tag == best_match and m.page_number == page_num + 1
                for m in matches
            )
            if already_matched_tag:
                continue

            color = _BBOX_COLORS[color_idx % len(_BBOX_COLORS)]
            color_idx += 1
            self._draw_bbox(page, x0, y0, x1, y1, best_match, color)
            matches.append(SiemensMatch(
                pdf_filename=pdf_filename,
                page_number=page_num + 1,
                tag=best_match,
                matched_text=word_text,
                bbox=(x0, y0, x1, y1),
                match_type="fuzzy",
                similarity_score=best_score,
                color_index=(color_idx - 1) % len(_BBOX_COLORS),
            ))
            already_matched_words.add(word_text)

        return matches

    # ── Bounding box drawing ───────────────────────────────────────────────

    def _draw_bbox(
        self,
        page: "fitz.Page",
        x0: float,
        y0: float,
        x1: float,
        y1: float,
        label: str,
        color: Tuple[float, float, float],
    ) -> None:
        """Draw a colored vector bounding box + label on the PDF page."""
        pad = 1.0
        rect = fitz.Rect(x0 - pad, y0 - pad, x1 + pad, y1 + pad)

        # Draw the rectangle (stroke only — keeps PDF text selectable)
        page.draw_rect(
            rect,
            color=color,
            width=_BBOX_STROKE_WIDTH,
            fill=None,
            overlay=True,
        )

        # Draw label background + text
        label_w = fitz.get_text_length(label, fontname="helv", fontsize=_LABEL_FONT_SIZE)
        label_h = _LABEL_FONT_SIZE + 2
        label_x = x0
        label_y = max(0, y0 - label_h - 1)
        label_rect = fitz.Rect(
            label_x,
            label_y,
            label_x + label_w + 4,
            label_y + label_h,
        )

        page.draw_rect(
            label_rect,
            color=color,
            fill=(1, 1, 1),
            width=0.5,
            overlay=True,
        )

        page.insert_textbox(
            label_rect,
            label,
            fontname="helv",
            fontsize=_LABEL_FONT_SIZE,
            color=color,
            align=fitz.TEXT_ALIGN_LEFT,
            overlay=True,
        )

    # ── Excel report ───────────────────────────────────────────────────────

    def _write_match_report(self, output_excel_path: str, result: SiemensResult) -> None:
        wb = Workbook()

        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Control IntelliCraft — Siemens Mode Match Report (v2)"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Generated:", datetime.now().isoformat(timespec="seconds")])
        ws.append(["IO List:", result.excel_path])
        ws.append(["PDFs processed:", len(result.pdf_paths)])
        ws.append([])
        ws.append(["Total IO tags:", result.total_io_tags])
        ws.append(["Matched tags:", len(result.matched_tags)])
        ws.append(["Unmatched tags:", len(result.unmatched_tags)])
        match_rate = (
            (len(result.matched_tags) / result.total_io_tags * 100)
            if result.total_io_tags else 0.0
        )
        ws.append(["Match rate (%):", round(match_rate, 2)])
        ws.append(["Total matches (incl. duplicates):", len(result.matches)])
        ws.append(["Duration (s):", round(result.duration_seconds, 2)])
        ws.append([])
        ws.append(["Match Type Breakdown:"])
        breakdown = result._match_type_breakdown()
        for mtype, count in sorted(breakdown.items()):
            ws.append([f"  {mtype}:", count])

        # Sheet 2: Matches
        ws2 = wb.create_sheet("Matches")
        headers = ["PDF File", "Page #", "Tag (IO List)", "Matched Text (PDF)",
                    "Match Type", "Similarity", "BBox x0", "BBox y0", "BBox x1", "BBox y1"]
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A3A6E")

        for m in result.matches:
            ws2.append([
                m.pdf_filename, m.page_number, m.tag, m.matched_text,
                m.match_type, round(m.similarity_score, 3),
                round(m.bbox[0], 2), round(m.bbox[1], 2),
                round(m.bbox[2], 2), round(m.bbox[3], 2),
            ])

        for col_letter, width in zip("ABCDEFGHIJ", [30, 8, 18, 22, 14, 10, 10, 10, 10, 10]):
            ws2.column_dimensions[col_letter].width = width

        # Sheet 3: Unmatched
        ws3 = wb.create_sheet("Unmatched")
        ws3.append(["Tag (IO List)"])
        ws3["A1"].font = Font(bold=True, color="FFFFFF")
        ws3["A1"].fill = PatternFill("solid", fgColor="C0392B")
        for tag in sorted(result.unmatched_tags):
            ws3.append([tag])
        ws3.column_dimensions["A"].width = 30

        # Sheet 4: Per PDF
        ws4 = wb.create_sheet("Per PDF")
        ws4.append(["PDF File", "Status", "Pages Processed", "Total Pages",
                     "Total Matches", "Matched Tags", "Pages with Matches"])
        for cell in ws4[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A3A6E")

        for pdf_name, stats in result.per_pdf.items():
            ws4.append([
                pdf_name, stats.get("status", "?"),
                stats.get("pages_processed", 0), stats.get("total_pages", 0),
                stats.get("total_matches", 0), stats.get("matched_tags_count", 0),
                ", ".join(str(p) for p in stats.get("pages_with_matches", [])),
            ])

        wb.save(output_excel_path)

    def _create_zip_bundle(
        self,
        output_pdf_dir: str,
        annotated_pdfs: List[str],
        report_excel: Optional[str],
        zip_path: Optional[str],
    ) -> Optional[str]:
        import zipfile
        if not annotated_pdfs and not report_excel:
            return None
        if zip_path is None:
            zip_path = output_pdf_dir.rstrip("/") + ".zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for pdf in annotated_pdfs:
                if os.path.exists(pdf):
                    zf.write(pdf, arcname=os.path.basename(pdf))
            if report_excel and os.path.exists(report_excel):
                zf.write(report_excel, arcname=os.path.basename(report_excel))
        return zip_path


# ── Convenience function ──────────────────────────────────────────────────
def run_siemens_mode(
    pdf_paths: List[str],
    excel_path: str,
    output_pdf_dir: str,
    output_excel_path: str,
    create_zip: bool = True,
    zip_path: Optional[str] = None,
) -> Dict[str, Any]:
    proc = SiemensModeProcessor()
    result = proc.process(
        pdf_paths=pdf_paths,
        excel_path=excel_path,
        output_pdf_dir=output_pdf_dir,
        output_excel_path=output_excel_path,
        create_zip=create_zip,
        zip_path=zip_path,
    )
    return result.to_dict()


# ── CLI ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse
    import json
    import sys

    parser = argparse.ArgumentParser(
        description="Siemens mode v2: find IO List tags in digital PDFs using DataAnalysisModule."
    )
    parser.add_argument("--pdf", action="append", required=True)
    parser.add_argument("--excel", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--out-excel", required=True)
    parser.add_argument("--no-zip", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO,
                         format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")

    result = run_siemens_mode(
        pdf_paths=args.pdf,
        excel_path=args.excel,
        output_pdf_dir=args.out_dir,
        output_excel_path=args.out_excel,
        create_zip=not args.no_zip,
    )
    print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
    sys.exit(0 if result.get("matched_count", 0) > 0 else 1)
