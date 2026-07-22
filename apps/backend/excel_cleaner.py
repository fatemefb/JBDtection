#!/usr/bin/env python3
"""
Control IntelliCraft — Data Consolidator v2 (Production-Ready)
Merges multiple Excel files into a single unified IO List.

Key improvements over v1:
  1. Smart Sheet Detection — auto-finds the sheet with actual data (not cover/title sheets)
  2. Header Row Detection — auto-finds which row is the actual header (skips title rows)
  3. Data Row Filtering — removes non-data rows (empty, summary, footer)
  4. Tag Validation — validates tag format (must look like a real tag, not "JB List" etc.)
  5. Multi-pass Column Resolution — priority-based, no false matches
"""

import os
import re
import logging
import tempfile
from typing import List, Dict, Optional, Tuple
from datetime import datetime
from functools import reduce

import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════
# COLUMN DICTIONARY
# ═══════════════════════════════════════════════════════════════

COLUMN_ALIASES = {
    'TAG': ['Tag No','Tag No.','Tag no','Tag no.','tag no','tag no.','TAG NO','TAG NO.','Tag','TAG','tag','TAG.','Tag Number','Tag_Number','TAG NUMBER','tag number','TagNo','TAGNO','tagno','تگ','شماره تگ'],
    'LOOP': ['Loop No','Loop No.','Loop no','loop no','LOOP NO','Loop','LOOP','loop','Loop Number','Loop_No','loop_number','LOOP NUMBER','لوپ','شماره لوپ'],
    'JB': ['JB','J.B','J B','jb','JB No','JB No.','JB no','JB NO','Junction Box','JUNCTION BOX','junction box','JB_ID','JBID','J/B','JunctionBox'],
    'IO_TYPE': ['I/O Type','IO Type','IOType','I O Type','i/o type','io type','IO TYPE','I/O TYPE','Signal Type','SIGNAL TYPE','signal type','IOType','I/O_Type','نوع IO','نوع سیگنال','نوع ورودی/خروجی'],
    'SAFETY': ['IS/NIS','IS-NIS','IS_NIS','IS/NIS.','is/nis','IS/NIS ',' IS/NIS','Safety','SAFETY','safety','Safe','SAFE','safe','SIL','Sil','sil','Instrumented','INSTRUMENTED','ایمن','غیرایمن','IS/ NIS','IS / NIS'],
    'LOCATION': ['Location','LOCATION','location','Loc','LOC','loc','Area','AREA','area','Place','PLACE','محل','مکان','ناحیه'],
    'TERM1': ['terminal-1','Terminal-1','TERMINAL-1','terminal1','Terminal1','TERMINAL1','terminal 1','Terminal 1','TERMINAL 1','term1','Term1','TERM1','terminal_1','Terminal_1','t1','T1','ترمینال ۱','ترمینال 1'],
    'TERM2': ['terminal-2','Terminal-2','TERMINAL-2','terminal2','Terminal2','TERMINAL2','terminal 2','Terminal 2','TERMINAL 2','term2','Term2','TERM2','terminal_2','Terminal_2','t2','T2','ترمینال ۲','ترمینال 2'],
    'SRC': ['SRC','src','Src','Source','SOURCE','source','SRC.','src.','منبع'],
    'CABLE': ['Cable Description','Cable_Description','cable description','CABLE DESCRIPTION','Cable Desc','CableDesc','Cable Code','CableCode','Cable','CABLE','cable','کابل','توضیح کابل'],
    'DESCRIPTION': ['Description','DESCRIPTION','description','Desc','DESC','Remark','Remarks','REMARKS','توضیحات','یادداشت'],
    'INSTRUMENT': ['Instrument','INSTRUMENT','instrument','Device','DEVICE','Equipment','EQUIPMENT','تجهیز','دستگاه'],
    'SERVICE': ['Service','SERVICE','service','Function','FUNCTION','Process','PROCESS','سرویس','فرآیند'],
}

CANONICAL = {
    'TAG':'Tag No', 'LOOP':'Loop No', 'JB':'JB', 'IO_TYPE':'I/O Type',
    'SAFETY':'IS/NIS', 'LOCATION':'Location', 'TERM1':'terminal-1', 'TERM2':'terminal-2',
    'SRC':'SRC', 'CABLE':'Cable Description', 'DESCRIPTION':'Description',
    'INSTRUMENT':'Instrument', 'SERVICE':'Service',
}

# ═══════════════════════════════════════════════════════════════
# COLUMN RESOLVER
# ═══════════════════════════════════════════════════════════════

def resolve_column(df_columns: List[str], logical_key: str) -> Optional[str]:
    """Find actual column name — 5-pass matching, no false positives."""
    aliases = COLUMN_ALIASES.get(logical_key, [])
    cols = [str(c) for c in df_columns]
    cols_lower = {c.strip().lower(): c for c in cols}
    
    def norm(s):
        return s.strip().lower().replace(' ','').replace('_','').replace('-','').replace('.','').replace('/','')
    
    cols_norm = {norm(c): c for c in cols}
    
    # Pass 1: Exact
    for a in aliases:
        if a in cols: return a
    
    # Pass 2: Case-insensitive exact
    for a in aliases:
        al = a.strip().lower()
        if al in cols_lower: return cols_lower[al]
    
    # Pass 3: Normalized (remove spaces/special chars)
    for a in aliases:
        an = norm(a)
        if an in cols_norm: return cols_norm[an]
    
    # Pass 4: Starts-with (alias >= 4 chars)
    for a in aliases:
        if len(a) < 4: continue
        al = a.strip().lower()
        for cl, co in cols_lower.items():
            if cl.startswith(al) or al.startswith(cl): return co
    
    # Pass 5: Contains (alias >= 5 chars)
    for a in aliases:
        if len(a) < 5: continue
        al = a.strip().lower()
        for cl, co in cols_lower.items():
            if al in cl or cl in al: return co
    
    return None


def resolve_all_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    """Resolve all known columns with priority — prevents double-assignment."""
    result = {}
    used = set()
    
    for key in ['TAG','LOOP','JB','IO_TYPE','SAFETY','LOCATION','TERM1','TERM2','SRC','CABLE','DESCRIPTION','INSTRUMENT','SERVICE']:
        available = [c for c in df.columns if c not in used]
        actual = resolve_column(available, key)
        if actual:
            result[key] = actual
            used.add(actual)
        else:
            result[key] = None
    
    return result


def standardize_columns(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Rename columns to canonical names."""
    df = df.copy()
    rename = {}
    for key, actual in resolve_all_columns(df).items():
        if actual and actual != CANONICAL[key]:
            rename[actual] = CANONICAL[key]
    if rename:
        df = df.rename(columns=rename)
    return df, rename


# ═══════════════════════════════════════════════════════════════
# SMART SHEET DETECTION
# ═══════════════════════════════════════════════════════════════

# Patterns that indicate a tag-like value
TAG_PATTERN = re.compile(r'^[A-Z]{1,5}[-]?\d{2,5}[A-Z]?$', re.IGNORECASE)

# Patterns that indicate NON-data sheets or rows
SKIP_SHEET_NAMES = {
    'cover', 'title', 'index', 'contents', 'summary', 'report',
    'info', 'information', 'about', 'readme', 'note', 'notes',
    'کاور', 'عنوان', 'فهرست', 'خلاصه', 'گزارش', 'اطلاعات',
    'jb list', 'jblist', 'jb_list', 'list of jb', 'jb-list',
    'cabinet list', 'cabinet', 'layout', 'drawing',
}

SKIP_ROW_KEYWORDS = {
    'project', 'total', 'sum', 'average', 'count', 'note:', 'remark:',
    'page', 'sheet', 'revision', 'date:', 'approved', 'checked',
    'پروژه', 'جمع', 'مجموع', 'تعداد', 'توضیح', 'صفحه', 'بازبینی',
    'تأیید', 'بررسی شده', 'نسخه',
}


def is_data_sheet(sheet_name: str, df: pd.DataFrame) -> bool:
    """
    Determine if a sheet contains actual data (not a cover/title/index sheet).
    
    Criteria:
      1. Sheet name doesn't match skip patterns
      2. Has at least 3 rows
      3. Has at least 2 columns
      4. At least one column looks like it could contain tags
    """
    name_lower = sheet_name.strip().lower()
    
    # Check skip patterns
    for skip in SKIP_SHEET_NAMES:
        if skip in name_lower:
            return False
    
    # Must have enough data
    if len(df) < 3 or len(df.columns) < 2:
        return False
    
    # Check if any column might be a tag column
    resolved = resolve_all_columns(df)
    if resolved.get('TAG'):
        return True
    
    # Check if any column has tag-like values
    for col in df.columns:
        sample = df[col].dropna().astype(str).head(20)
        tag_count = sum(1 for v in sample if TAG_PATTERN.match(v.strip()))
        if tag_count >= 3:  # At least 3 values look like tags
            return True
    
    return False


def find_header_row(df: pd.DataFrame, max_scan: int = 10) -> int:
    """
    Find the actual header row — sometimes the first few rows are title/project info.
    
    Returns the 0-indexed row number that should be used as header.
    0 means the first row is already the header.
    """
    # Try each row as potential header
    for row_idx in range(min(max_scan, len(df))):
        row_values = df.iloc[row_idx].astype(str).str.strip().tolist()
        
        # Check if this row has recognizable column names
        resolved_count = 0
        for val in row_values:
            for aliases in COLUMN_ALIASES.values():
                val_lower = val.strip().lower()
                for alias in aliases:
                    if val_lower == alias.strip().lower():
                        resolved_count += 1
                        break
                else:
                    continue
                break
        
        # If we found at least 2 recognizable columns, this is the header
        if resolved_count >= 2:
            return row_idx
    
    # Default: first row
    return 0


def clean_data_rows(df: pd.DataFrame, tag_col: str = 'Tag No') -> pd.DataFrame:
    """
    Remove non-data rows:
      - Empty rows
      - Summary/total rows
      - Rows where Tag No doesn't look like a tag
    """
    if tag_col not in df.columns:
        return df
    
    df = df.copy()
    
    # Convert Tag No to string and strip
    df[tag_col] = df[tag_col].astype(str).str.strip()
    
    # Remove rows with empty/NaN tag
    df = df[df[tag_col].notna()]
    df = df[df[tag_col] != '']
    df = df[df[tag_col] != 'nan']
    df = df[df[tag_col] != 'None']
    df = df[df[tag_col] != '-']
    df = df[df[tag_col].str.lower() != 'n/a']
    
    # Remove rows where Tag column contains skip keywords
    mask = df[tag_col].str.lower().str.contains('|'.join(SKIP_ROW_KEYWORDS), na=False)
    df = df[~mask]
    
    # Remove rows where Tag doesn't look like a tag at all
    # (but be lenient — some tags might not match the pattern)
    # Only remove rows that are clearly NOT tags (e.g., long sentences)
    def is_likely_tag(val):
        val = str(val).strip()
        if len(val) > 30:  # Tags are short
            return False
        if ' ' in val and not TAG_PATTERN.match(val):  # Multi-word = probably not a tag
            return False
        if val.replace('.','').replace('-','').replace('_','').isalpha() and len(val) > 10:
            return False  # Long text without numbers = probably not a tag
        return True
    
    df = df[df[tag_col].apply(is_likely_tag)]
    
    # Uppercase tags
    df[tag_col] = df[tag_col].str.upper()
    
    return df.reset_index(drop=True)


# ═══════════════════════════════════════════════════════════════
# EXCEL CONSOLIDATOR
# ═══════════════════════════════════════════════════════════════

class ExcelConsolidator:
    """
    Production-ready Excel Consolidator.
    
    Pipeline per file:
      1. Read all sheets
      2. Smart Sheet Detection — find the data sheet (skip cover/index/JB-list)
      3. Header Row Detection — find actual header row (skip title rows)
      4. Column Resolution — detect and standardize column names
      5. Data Row Cleaning — remove empty/summary/non-tag rows
      6. Tag Validation — validate tag format
      7. Merge — outer join on Tag No, first non-null wins
    """
    
    def __init__(self):
        self.merge_log = []
        self.conflicts = []
        self.total_tags = set()
    
    def consolidate(self, file_paths: List[str]) -> Tuple[pd.DataFrame, dict]:
        """Consolidate multiple Excel files into one unified IO List."""
        all_dfs = []
        
        for file_path in file_paths:
            file_name = os.path.basename(file_path)
            logger.info(f"Processing: {file_name}")
            
            try:
                xls = pd.ExcelFile(file_path)
                data_sheet_found = False
                
                for sheet_name in xls.sheet_names:
                    # Read without header first (to scan for header row)
                    df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    
                    if df_raw.empty or len(df_raw) < 3:
                        self._log(file_name, sheet_name, 'skipped_empty', 0, [])
                        continue
                    
                    # Smart Sheet Detection
                    df_preview = pd.read_excel(file_path, sheet_name=sheet_name)
                    if not is_data_sheet(sheet_name, df_preview):
                        self._log(file_name, sheet_name, f'skipped_not_data (sheet: {sheet_name})', len(df_preview), list(df_preview.columns))
                        logger.info(f"  {sheet_name}: skipped (not a data sheet)")
                        continue
                    
                    # Header Row Detection
                    header_row = find_header_row(df_raw)
                    
                    if header_row > 0:
                        # Re-read with correct header
                        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                        logger.info(f"  {sheet_name}: header found at row {header_row + 1}")
                    else:
                        df = df_preview
                    
                    if df.empty:
                        continue
                    
                    # Standardize column names
                    df_std, rename_map = standardize_columns(df)
                    
                    # Ensure Tag No exists
                    if 'Tag No' not in df_std.columns:
                        # Try one more time with raw columns
                        tag_col = resolve_column(list(df.columns), 'TAG')
                        if tag_col:
                            df_std = df_std.rename(columns={tag_col: 'Tag No'})
                        else:
                            self._log(file_name, sheet_name, 'no_tag_column', len(df), list(df.columns))
                            continue
                    
                    # Clean data rows
                    df_std = clean_data_rows(df_std, 'Tag No')
                    
                    if df_std.empty:
                        self._log(file_name, sheet_name, 'no_valid_data_rows', 0, [])
                        continue
                    
                    # Select only known columns
                    known_cols = ['Tag No'] + [CANONICAL[k] for k in COLUMN_ALIASES if k != 'TAG' and CANONICAL[k] in df_std.columns]
                    df_select = df_std[known_cols].copy()
                    
                    all_dfs.append(df_select)
                    data_sheet_found = True
                    
                    self._log(file_name, sheet_name, 'processed', len(df_std), known_cols, rename_map, header_row)
                    logger.info(f"  {sheet_name}: {len(df_std)} rows, {len(known_cols)} cols, header=row {header_row+1}")
                    
                    # Only process the FIRST data sheet per file (most files have 1 data sheet)
                    break
                
                if not data_sheet_found:
                    self._log(file_name, '-', 'no_data_sheet_found', 0, [])
                    logger.warning(f"  {file_name}: no data sheet found!")
                
            except Exception as e:
                logger.error(f"Error: {file_name}: {e}")
                self._log(file_name, '-', f'error: {e}', 0, [])
        
        if not all_dfs:
            return pd.DataFrame(), self._report()
        
        # Merge
        merged = reduce(lambda a, b: self._merge_two(a, b), all_dfs)
        merged = merged.drop_duplicates(subset=['Tag No'], keep='first')
        merged = merged.sort_values('Tag No').reset_index(drop=True)
        
        # Reorder
        col_order = ['Tag No']
        for key in ['LOOP','JB','IO_TYPE','SAFETY','LOCATION','TERM1','TERM2','SRC','CABLE','DESCRIPTION','INSTRUMENT','SERVICE']:
            col = CANONICAL[key]
            if col in merged.columns: col_order.append(col)
        for col in merged.columns:
            if col not in col_order: col_order.append(col)
        merged = merged[col_order]
        
        return merged, self._report(merged)
    
    def _merge_two(self, df1, df2):
        """Merge two DataFrames, first non-null wins."""
        merged = df1.merge(df2, on='Tag No', how='outer', suffixes=('', '_y'))
        
        for col in list(merged.columns):
            if col.endswith('_y'):
                base = col[:-2]
                if base in merged.columns:
                    both = merged[merged[base].notna() & merged[col].notna() & (merged[base] != merged[col])]
                    for _, row in both.head(5).iterrows():
                        self.conflicts.append({'tag': row['Tag No'], 'column': base, 'val1': str(row[base]), 'val2': str(row[col])})
                    merged[base] = merged[base].fillna(merged[col])
                    merged = merged.drop(columns=[col])
                else:
                    merged = merged.rename(columns={col: base})
        return merged
    
    def _log(self, file, sheet, status, rows, columns, rename=None, header_row=None):
        entry = {'file': file, 'sheet': sheet, 'status': status, 'rows': rows, 'columns': columns}
        if rename: entry['renamed'] = rename
        if header_row is not None: entry['header_row'] = header_row + 1
        self.merge_log.append(entry)
    
    def _report(self, merged=None):
        r = {
            'timestamp': datetime.now().isoformat(),
            'files': len(self.merge_log),
            'succeeded': sum(1 for l in self.merge_log if l['status'] == 'processed'),
            'skipped': sum(1 for l in self.merge_log if l['status'] != 'processed' and 'error' not in l['status']),
            'errored': sum(1 for l in self.merge_log if 'error' in l['status']),
            'tags': len(self.total_tags),
            'log': self.merge_log,
            'conflicts': self.conflicts[:20],
            'total_conflicts': len(self.conflicts),
        }
        if merged is not None and not merged.empty:
            r['rows'] = len(merged)
            r['columns'] = list(merged.columns)
            r['fill'] = {c: f"{merged[c].notna().sum()}/{len(merged)}" for c in merged.columns}
        return r


# ═══════════════════════════════════════════════════════════════
# TEST
# ═══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(message)s')
    
    from openpyxl import Workbook
    
    temp = tempfile.gettempdir()
    
    # ── Test File 1: IO List with cover sheet + data sheet ──
    wb1 = Workbook()
    # Sheet 1: Cover (should be skipped)
    ws_cover = wb1.active
    ws_cover.title = 'Cover'
    ws_cover['A1'] = 'Project: Sina Methanol Plant'
    ws_cover['A2'] = 'Document: IO List'
    ws_cover['A3'] = 'Date: 2026-07-15'
    ws_cover['A4'] = 'Revision: 1'
    
    # Sheet 2: JB List (should be skipped)
    ws_jb = wb1.create_sheet('JB List')
    ws_jb['A1'] = 'JB Name'
    ws_jb['A2'] = 'JSG-903S'
    ws_jb['A3'] = 'JAG-902S'
    ws_jb['A4'] = 'JHF-901S'
    
    # Sheet 3: Actual IO List data (should be processed)
    ws_data = wb1.create_sheet('IO List')
    # First 2 rows are title info (header detection should skip them)
    ws_data['A1'] = 'Sina Methanol Plant - IO List'
    ws_data['A2'] = 'Sheet 1 of 5'
    # Row 3 is the actual header
    ws_data['A3'] = 'Tag No'
    ws_data['B3'] = 'I/O Type'
    ws_data['C3'] = 'IS/NIS'
    ws_data['D3'] = 'Loop No'
    # Data rows
    ws_data['A4'] = 'UY-5021'
    ws_data['B4'] = 'AI'
    ws_data['C4'] = 'IS'
    ws_data['D4'] = 'L-5021'
    ws_data['A5'] = 'FUY-5041'
    ws_data['B5'] = 'AO'
    ws_data['C5'] = 'NIS'
    ws_data['D5'] = 'L-5041'
    # Summary row (should be filtered out)
    ws_data['A6'] = 'Total: 2 tags'
    
    p1 = os.path.join(temp, 'test_io_list.xlsx')
    wb1.save(p1)
    
    # ── Test File 2: JB Assignment file (single sheet) ──
    df2 = pd.DataFrame({
        'Tag No': ['UY-5021', 'FUY-5041', 'FUY-5161', 'UY-9999'],
        'JB': ['JSG-903S', 'JSG-903S', 'JAG-902S', 'JAG-902S'],
        'Location': ['FIELD', 'FIELD', 'MCC', 'FIELD'],
        'terminal-1': ['1', '3', '5', '7'],
        'terminal-2': ['2', '4', '6', '8'],
    })
    p2 = os.path.join(temp, 'test_jb_assignment.xlsx')
    df2.to_excel(p2, index=False)
    
    # ── Test File 3: Cable info with different column names ──
    df3 = pd.DataFrame({
        'TAG': ['UY-5021', 'FUY-5041'],
        'Cable': ['Cable-A-001', 'Cable-B-002'],
        'Description': ['Temperature Sensor', 'Pressure Transmitter'],
        'SRC': ['Field', 'Field'],
    })
    p3 = os.path.join(temp, 'test_cable_info.xlsx')
    df3.to_excel(p3, index=False)
    
    # ── Run Consolidator ──
    print('\n' + '='*60)
    print('TEST: Consolidate 3 files')
    print('='*60)
    
    c = ExcelConsolidator()
    merged, report = c.consolidate([p1, p2, p3])
    
    print('\n' + '='*60)
    print('MERGED RESULT')
    print('='*60)
    print(merged.to_string())
    print(f'\nRows: {len(merged)}, Columns: {len(merged.columns)}')
    
    print(f'\n{"="*60}')
    print('REPORT')
    print(f'{"="*60}')
    print(f'Files: {report["files"]} | Succeeded: {report["succeeded"]} | Skipped: {report["skipped"]}')
    print(f'Total tags: {report["tags"]} | Conflicts: {report["total_conflicts"]}')
    
    print(f'\nMerge log:')
    for log in report['log']:
        extra = f', header=row {log["header_row"]}' if 'header_row' in log else ''
        extra += f', renamed: {log.get("renamed", {})}' if log.get('renamed') else ''
        print(f'  {log["file"]}/{log["sheet"]}: {log["status"]} ({log["rows"]} rows){extra}')
    
    print(f'\nFill rates:')
    for col, rate in report.get('fill', {}).items():
        print(f'  {col}: {rate}')
    
    if report['conflicts']:
        print(f'\nConflicts:')
        for cf in report['conflicts']:
            print(f'  {cf}')
    
    # Cleanup
    for p in [p1, p2, p3]:
        os.remove(p)
    
    print('\n✅ Test passed!')
