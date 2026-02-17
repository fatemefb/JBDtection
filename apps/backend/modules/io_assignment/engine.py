import math
import re
import logging
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


logger = logging.getLogger(__name__)

DEFAULT_CHANNELS = {
    "Barrier Board (AI)": 16,
    "Barrier Board (AO)": 16,
    "Barrier Board (DI)": 16,
    "Barrier Board (DO)": 16,
    "Terminal Board (AI)": 16,
    "Terminal Board (AO)": 16,
    "Terminal Board (DI)": 16,
    "Terminal Board (DO)": 16,
    "Relay Board AO capacity": 32,
    "Relay Board AI capacity": 32,
    "Relay Board DO capacity": 32,
    "Relay Board DI capacity": 32,
}

DEFAULT_CABINET_PLAN = [
    {"type": "Type 1", "direction": "REAR", "quantity": 2, "priority": 1},
    {"type": "Type 1", "direction": "FRONT", "quantity": 2, "priority": 2},
    {"type": "Type 2", "direction": "REAR", "quantity": 2, "priority": 3},
    {"type": "Type 2", "direction": "FRONT", "quantity": 2, "priority": 4},
    {"type": "Type 3", "direction": "REAR", "quantity": 2, "priority": 5},
    {"type": "Type 3", "direction": "FRONT", "quantity": 2, "priority": 6},
]

DEFAULT_CABINET_LIMITS = {
    ("Type 1", "REAR"): {
        "Max_Total_Boards": 100,
        "Max rail terminals": 800,
        "Barrier Board (AI)": 20,
        "Barrier Board (AO)": 20,
        "Barrier Board (DI)": 12.5,
        "Barrier Board (DO)": 12.5,
        "Terminal Board (AI)": 20,
        "Terminal Board (AO)": 20,
        "Terminal Board (DI)": 12.5,
        "Terminal Board (DO)": 12.5,
        "Relay Board AO capacity": 25,
        "Relay Board AI capacity": 25,
        "Relay Board DO capacity": 33.3,
        "Relay Board DI capacity": 33.3,
    },
    ("Type 1", "FRONT"): {
        "Max_Total_Boards": 100,
        "Max rail terminals": 800,
        "Barrier Board (AI)": 14.2,
        "Barrier Board (AO)": 14.2,
        "Barrier Board (DI)": 10,
        "Barrier Board (DO)": 10,
        "Terminal Board (AI)": 14.2,
        "Terminal Board (AO)": 14.2,
        "Terminal Board (DI)": 7.1,
        "Terminal Board (DO)": 8.3,
        "Relay Board AO capacity": 14.2,
        "Relay Board AI capacity": 14.2,
        "Relay Board DO capacity": 16.6,
        "Relay Board DI capacity": 16.6,
    },
    ("Type 2", "FRONT"): {
        "Max_Total_Boards": 100,
        "Max rail terminals": 800,
        "Barrier Board (AI)": 14.2,
        "Barrier Board (AO)": 14.2,
        "Barrier Board (DI)": 10,
        "Barrier Board (DO)": 10,
        "Terminal Board (AI)": 14.2,
        "Terminal Board (AO)": 14.2,
        "Terminal Board (DI)": 7.1,
        "Terminal Board (DO)": 8.3,
        "Relay Board AO capacity": 14.2,
        "Relay Board AI capacity": 14.2,
        "Relay Board DO capacity": 16.6,
        "Relay Board DI capacity": 16.6,
    },
    ("Type 2", "REAR"): {
        "Max_Total_Boards": 100,
        "Max rail terminals": 800,
        "Barrier Board (AI)": 14.2,
        "Barrier Board (AO)": 14.2,
        "Barrier Board (DI)": 20,
        "Barrier Board (DO)": 20,
        "Terminal Board (AI)": 14.2,
        "Terminal Board (AO)": 14.2,
        "Terminal Board (DI)": 7.1,
        "Terminal Board (DO)": 8.3,
        "Relay Board AO capacity": 14.2,
        "Relay Board AI capacity": 14.2,
        "Relay Board DO capacity": 16.6,
        "Relay Board DI capacity": 16.6,
    },
    ("Type 3", "REAR"): {
        "Max_Total_Boards": 100,
        "Max rail terminals": 800,
        "Barrier Board (AI)": 16.6,
        "Barrier Board (AO)": 16.6,
        "Barrier Board (DI)": 10,
        "Barrier Board (DO)": 10,
        "Terminal Board (AI)": 20,
        "Terminal Board (AO)": 20,
        "Terminal Board (DI)": 10,
        "Terminal Board (DO)": 10,
        "Relay Board AO capacity": 20,
        "Relay Board AI capacity": 20,
        "Relay Board DO capacity": 25,
        "Relay Board DI capacity": 25,
    },
    ("Type 3", "FRONT"): {
        "Max_Total_Boards": 100,
        "Max rail terminals": 800,
        "Barrier Board (AI)": 14.2,
        "Barrier Board (AO)": 14.2,
        "Barrier Board (DI)": 20,
        "Barrier Board (DO)": 20,
        "Terminal Board (AI)": 14.2,
        "Terminal Board (AO)": 14.2,
        "Terminal Board (DI)": 7.1,
        "Terminal Board (DO)": 8.3,
        "Relay Board AO capacity": 14.2,
        "Relay Board AI capacity": 14.2,
        "Relay Board DO capacity": 16.6,
        "Relay Board DI capacity": 16.6,
    },
}


@dataclass
class IOConfig:
    channels: Dict[str, float] = field(default_factory=lambda: dict(DEFAULT_CHANNELS))
    hot_spare_ratio: float = 0.20
    jb_label_overhead: int = 2
    is_cab_prefix: str = "IS-A-"
    nis_cab_prefix: str = "NIS-B-"
    has_directions: bool = True
    rank_io: Dict[str, int] = field(default_factory=lambda: {
        "AI": 1, "AIR": 2, "AOR": 3, "AO": 4,
        "DI": 5, "DIR": 6, "DO": 7, "DOR": 8,
    })
    rank_safety: Dict[str, int] = field(default_factory=lambda: {"IS": 1, "NIS": 2})
    rank_location: Dict[str, int] = field(default_factory=lambda: {"FIELD": 1, "MCC": 2, "LCP": 3, "UCP": 4})
    col_mapping: Dict[str, str] = field(default_factory=lambda: {
        "JB": "JB",
        "IO_TYPE": "I/O Type",
        "SAFETY": "IS/NIS",
        "LOCATION": "Location",
        "TERM1": "terminal-1",
        "TERM2": "terminal-2",
        "SRC": "SRC",
        "SPARE_COUNT": "JB_SPARE_COUNT",
    })
    io_normal: Dict[str, str] = field(default_factory=lambda: {
        "AI": "AI", "A.I": "AI", "A I": "AI",
        "AO": "AO", "A.O": "AO", "A O": "AO",
        "AIR": "AIR", "A.I.R": "AIR", "A I R": "AIR",
        "AOR": "AOR", "A.O.R": "AOR", "A O R": "AOR",
        "DI": "DI", "D.I": "DI", "D I": "DI",
        "DO": "DO", "D.O": "DO", "D O": "DO",
        "DIR": "DIR", "DI-R": "DIR",
        "DOR": "DOR", "DO-R": "DOR",
    })
    safety_normal: Dict[str, str] = field(default_factory=lambda: {
        "IS": "IS", "I.S": "IS",
        "NIS": "NIS", "N.I.S": "NIS",
    })
    location_normal: Dict[str, str] = field(default_factory=lambda: {
        "FIELD": "FIELD",
        "FIELD-AREA": "FIELD",
        "MCC": "MCC",
        "MCC ROOM": "MCC",
        "LCP": "LCP",
        "LOCAL": "LCP",
        "UCP": "UCP",
    })
    invalid_jb: set = field(default_factory=lambda: {"-", "N/A", "0", "UNKNOWN", "NAN", ""})
    cabinet_plan: List[Dict[str, Any]] = field(default_factory=lambda: list(DEFAULT_CABINET_PLAN))
    cabinet_limits: Dict[Tuple[str, str], Dict[str, float]] = field(
        default_factory=lambda: dict(DEFAULT_CABINET_LIMITS)
    )


def build_config(overrides: Optional[Dict[str, Any]]) -> IOConfig:
    config = IOConfig()
    if not overrides:
        return config

    if "channels" in overrides and isinstance(overrides["channels"], dict):
        config.channels.update(overrides["channels"])

    if "hot_spare_ratio" in overrides:
        config.hot_spare_ratio = float(overrides["hot_spare_ratio"])
    if "jb_label_overhead" in overrides:
        config.jb_label_overhead = int(overrides["jb_label_overhead"])
    if "has_directions" in overrides:
        config.has_directions = bool(overrides["has_directions"])

    if "cabinet_plan" in overrides and isinstance(overrides["cabinet_plan"], list):
        config.cabinet_plan = overrides["cabinet_plan"]

    if "cabinet_limits" in overrides and isinstance(overrides["cabinet_limits"], dict):
        limits = {}
        for key, val in overrides["cabinet_limits"].items():
            if isinstance(key, (list, tuple)) and len(key) == 2:
                limits[(key[0], key[1])] = val
            elif isinstance(key, str) and "|" in key:
                t, d = key.split("|", 1)
                limits[(t.strip(), d.strip())] = val
        if limits:
            config.cabinet_limits = limits

    return config


def build_cabinet_pool_from_plan(config: IOConfig) -> List[Dict[str, Any]]:
    pool = []
    plan_sorted = sorted(config.cabinet_plan, key=lambda x: x["priority"])
    for item in plan_sorted:
        t = item["type"]
        d = item["direction"]
        qty = int(item["quantity"])
        limits = config.cabinet_limits.get((t, d))
        if limits is None and d == "SINGLE":
            limits = config.cabinet_limits.get((t, "FRONT")) or config.cabinet_limits.get((t, "REAR"))
        if limits is None:
            raise ValueError(f"Missing limits for cabinet type/direction: {(t, d)}")
        for _ in range(qty):
            pool.append({
                "name": t,
                "direction": d,
                "limits": limits,
            })
    return pool


def normalize_io_for_map(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).upper().strip()
    if s in ("", "NAN", "NONE", "NULL", "-"):
        return ""
    s = s.replace("_", " ").replace("-", " ").replace(".", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if s in ("DI R", "D I R"):
        return "DI-R"
    if s in ("DO R", "D O R"):
        return "DO-R"
    return s


def initial_cleaning_and_report(df_raw: pd.DataFrame, config: IOConfig):
    report = {
        "Summary": {},
        "Removed_Missing_Critical": [],
        "Removed_Invalid_JB": [],
        "Invalid_IO": [],
        "Safety_Conversion": [],
        "Location_Imputed": [],
        "Logical_Exceptions": [],
    }

    df = df_raw.copy()
    total_raw = len(df)

    for col in df.select_dtypes(include=["object"]).columns:
        df[col] = (
            df[col]
            .astype(str)
            .str.upper()
            .str.strip()
            .replace({"NAN": "", "NONE": "", "?": ""})
        )

    jb_col = config.col_mapping["JB"]
    io_col = config.col_mapping["IO_TYPE"]
    saf_col = config.col_mapping["SAFETY"]
    loc_col = config.col_mapping["LOCATION"]

    src_col = config.col_mapping["SRC"]
    if src_col not in df.columns:
        df[src_col] = ""
    spare_count_col = config.col_mapping.get("SPARE_COUNT", "JB_SPARE_COUNT")
    input_has_spare_count = spare_count_col in df.columns
    if not input_has_spare_count:
        df[spare_count_col] = pd.NA
    df["_HAS_SPARE_COUNT_INPUT"] = bool(input_has_spare_count)
    df[spare_count_col] = pd.to_numeric(df[spare_count_col], errors="coerce").fillna(0).clip(lower=0)

    crit_mask = (df[jb_col] == "") | (df[io_col] == "")
    report["Removed_Missing_Critical"] = df[crit_mask].to_dict("records")
    df = df[~crit_mask]

    invalid_jb_mask = df[jb_col].isin(config.invalid_jb)
    report["Removed_Invalid_JB"] = df[invalid_jb_mask].to_dict("records")
    df = df[~invalid_jb_mask]

    df["IO_RAW"] = df[io_col]
    df["IO_KEY"] = df["IO_RAW"].apply(normalize_io_for_map)
    df[io_col] = df["IO_KEY"].map(config.io_normal)

    invalid_io = df[df[io_col].isna()]
    report["Invalid_IO"] = invalid_io[[jb_col, "IO_RAW", "IO_KEY"]].to_dict("records")
    df = df.dropna(subset=[io_col])

    df["SAFETY_RAW"] = df[saf_col]
    df[saf_col] = df["SAFETY_RAW"].map(config.safety_normal)
    safety_conv_mask = df["SAFETY_RAW"] != df[saf_col]
    report["Safety_Conversion"] = df[safety_conv_mask][["SAFETY_RAW", saf_col]].to_dict("records")
    df = df.dropna(subset=[saf_col])

    df["LOC_RAW"] = df[loc_col]
    df[loc_col] = df["LOC_RAW"].map(config.location_normal)
    loc_imp_mask = df[loc_col].isna()
    report["Location_Imputed"] = df[loc_imp_mask].to_dict("records")
    df.loc[loc_imp_mask, loc_col] = "FIELD"

    df["Signal_Type"] = "ACTIVE"
    df["Rank_IO"] = df[io_col].map(config.rank_io).fillna(99)
    df["Rank_Safety"] = df[saf_col].map(config.rank_safety).fillna(99)
    df["Rank_Loc"] = df[loc_col].map(config.rank_location).fillna(99)

    logic_mask = (df[io_col].isin(["AI", "AO"])) & (df[loc_col] == "MCC")
    report["Logical_Exceptions"] = df[logic_mask].to_dict("records")

    report["Summary"] = {
        "Total_Raw": total_raw,
        "Total_Clean": len(df),
        "Removed": total_raw - len(df),
    }

    return df.reset_index(drop=True), report


def master_sort(df: pd.DataFrame, config: IOConfig) -> pd.DataFrame:
    return df.sort_values(
        by=["Rank_IO", "Rank_Safety", "Rank_Loc", config.col_mapping["JB"]],
        ascending=[True, True, True, True],
    ).reset_index(drop=True)


def inject_hot_spares(df: pd.DataFrame, config: IOConfig) -> pd.DataFrame:
    spare_rows = []
    io_col = config.col_mapping["IO_TYPE"]
    saf_col = config.col_mapping["SAFETY"]
    jb_col = config.col_mapping["JB"]
    term1_col = config.col_mapping["TERM1"]
    term2_col = config.col_mapping["TERM2"]
    spare_count_col = config.col_mapping.get("SPARE_COUNT", "JB_SPARE_COUNT")

    def parse_terminal(value: Any) -> Optional[Tuple[str, int, int]]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        match = re.match(r"(.*?)(\d+)$", text)
        if not match:
            return None
        prefix, digits = match.group(1), match.group(2)
        return prefix, int(digits), len(digits)

    def format_terminal(prefix: str, width: int, number: int) -> str:
        return f"{prefix}{number:0{width}d}"

    terminal_state: Dict[str, Dict[str, Any]] = {}
    if term1_col in df.columns or term2_col in df.columns:
        for jb_name, group in df.groupby(jb_col):
            max_num = 0
            fmt1: Optional[Tuple[str, int]] = None
            fmt2: Optional[Tuple[str, int]] = None
            if term1_col in group.columns:
                for val in group[term1_col]:
                    parsed = parse_terminal(val)
                    if parsed:
                        prefix, number, width = parsed
                        max_num = max(max_num, number)
                        if fmt1 is None:
                            fmt1 = (prefix, width)
            if term2_col in group.columns:
                for val in group[term2_col]:
                    parsed = parse_terminal(val)
                    if parsed:
                        prefix, number, width = parsed
                        max_num = max(max_num, number)
                        if fmt2 is None:
                            fmt2 = (prefix, width)
            if fmt1 is None:
                fmt1 = ("T", 1)
            if fmt2 is None:
                fmt2 = ("T", 1)
            terminal_state[jb_name] = {
                "next": max_num + 1,
                "fmt1": fmt1,
                "fmt2": fmt2,
            }

    grouped = df.groupby([io_col, saf_col])

    has_spare_flag = bool(df.get("_HAS_SPARE_COUNT_INPUT", pd.Series([False])).astype(bool).any())
    enforce_spare_cap = spare_count_col in df.columns and has_spare_flag
    remaining_spare_by_jb: Dict[str, int] = {}
    if enforce_spare_cap:
        spare_capacity_df = df[[jb_col, spare_count_col]].copy()
        spare_capacity_df[jb_col] = spare_capacity_df[jb_col].astype(str).str.strip()
        spare_capacity_df[spare_count_col] = (
            pd.to_numeric(spare_capacity_df[spare_count_col], errors="coerce")
            .fillna(0)
            .clip(lower=0)
            .astype(int)
        )
        remaining_spare_by_jb = (
            spare_capacity_df.groupby(jb_col)[spare_count_col]
            .max()
            .to_dict()
        )

    for (io_type, safety), group in grouped:
        total_signals = len(group)
        spares_needed = math.ceil(total_signals * config.hot_spare_ratio)
        if spares_needed == 0:
            continue

        jbs = [str(jb).strip() for jb in group[jb_col].unique() if str(jb).strip()]
        if not jbs:
            continue

        allocated = 0
        max_rounds = (spares_needed * max(len(jbs), 1)) + 10
        rounds = 0

        while allocated < spares_needed and rounds < max_rounds:
            rounds += 1
            allocated_this_round = 0
            for jb in jbs:
                if allocated >= spares_needed:
                    break
                if enforce_spare_cap and remaining_spare_by_jb.get(jb, 0) <= 0:
                    continue

                rep_row = group[group[jb_col].astype(str).str.strip() == jb].iloc[0].copy()
                rep_row["Tag No"] = "HOT_SPARE"
                rep_row["Loop No"] = "HOT_SPARE"
                rep_row["Signal_Type"] = "HOT_SPARE"
                if jb in terminal_state and (term1_col in df.columns or term2_col in df.columns):
                    state = terminal_state[jb]
                    term1_prefix, term1_width = state["fmt1"]
                    term2_prefix, term2_width = state["fmt2"]
                    term1_num = state["next"]
                    term2_num = state["next"] + 1
                    state["next"] += 2
                    if term1_col in df.columns:
                        rep_row[term1_col] = format_terminal(term1_prefix, term1_width, term1_num)
                    if term2_col in df.columns:
                        rep_row[term2_col] = format_terminal(term2_prefix, term2_width, term2_num)
                spare_rows.append(rep_row)
                allocated += 1
                allocated_this_round += 1
                if enforce_spare_cap:
                    remaining_spare_by_jb[jb] = max(remaining_spare_by_jb.get(jb, 0) - 1, 0)

            if allocated_this_round == 0:
                break

        if enforce_spare_cap and allocated < spares_needed:
            logger.info(
                "HOT_SPARE capped by JB spare capacity for group (%s, %s): requested=%s allocated=%s",
                io_type,
                safety,
                spares_needed,
                allocated,
            )

    if spare_rows:
        df = pd.concat([df, pd.DataFrame(spare_rows)], ignore_index=True)

    if "_HAS_SPARE_COUNT_INPUT" in df.columns:
        df = df.drop(columns=["_HAS_SPARE_COUNT_INPUT"])

    return df


class JBDemand:
    def __init__(self, name: str, safety_type: str):
        self.name = name
        self.safety_type = safety_type
        self.signals: List[Dict[str, Any]] = []
        self.rail_load = 0
        self.channel_counts = {
            "Barrier_AI": 0,
            "Barrier_AO": 0,
            "Barrier_DI": 0,
            "Barrier_DO": 0,
            "Terminal_AI": 0,
            "Terminal_AO": 0,
            "Terminal_DI": 0,
            "Terminal_DO": 0,
            "Relay_DI": 0,
            "Relay_DO": 0,
            "Relay_AI": 0,
            "Relay_AO": 0,
        }


def calculate_demands(df: pd.DataFrame, config: IOConfig):
    jb_demands: Dict[str, JBDemand] = {}
    jb_col = config.col_mapping["JB"]
    io_col = config.col_mapping["IO_TYPE"]
    saf_col = config.col_mapping["SAFETY"]
    loc_col = config.col_mapping["LOCATION"]
    src_col = config.col_mapping["SRC"]

    for jb_name, group in df.groupby(jb_col):
        safety = group.iloc[0][saf_col]
        demand = JBDemand(jb_name, safety)
        demand.signals = group.to_dict("records")

        scr_signals = (group[src_col] == "SC").sum()
        demand.rail_load = (len(group) * 2) + scr_signals + config.jb_label_overhead

        for io_type, gi in group.groupby(io_col):
            cnt = len(gi)
            loc = gi.iloc[0][loc_col]

            base_io = io_type
            if io_type == "DIR":
                base_io = "DI"
            elif io_type == "DOR":
                base_io = "DO"
            elif io_type == "AIR":
                base_io = "AI"
            elif io_type == "AOR":
                base_io = "AO"

            if safety == "IS":
                if base_io == "AI":
                    demand.channel_counts["Barrier_AI"] += cnt
                elif base_io == "AO":
                    demand.channel_counts["Barrier_AO"] += cnt
                elif base_io == "DI":
                    demand.channel_counts["Barrier_DI"] += cnt
                elif base_io == "DO":
                    demand.channel_counts["Barrier_DO"] += cnt
            else:
                if base_io == "AI":
                    demand.channel_counts["Terminal_AI"] += cnt
                elif base_io == "AO":
                    demand.channel_counts["Terminal_AO"] += cnt
                elif base_io == "DI":
                    demand.channel_counts["Terminal_DI"] += cnt
                elif base_io == "DO":
                    demand.channel_counts["Terminal_DO"] += cnt

            if loc == "MCC" and base_io in ["DI", "DO", "AI", "AO"]:
                if base_io == "DI":
                    demand.channel_counts["Relay_DI"] += cnt
                elif base_io == "DO":
                    demand.channel_counts["Relay_DO"] += cnt
                elif base_io == "AI":
                    demand.channel_counts["Relay_AI"] += cnt
                elif base_io == "AO":
                    demand.channel_counts["Relay_AO"] += cnt

        jb_demands[jb_name] = demand

    return jb_demands


class Cabinet:
    def __init__(self, cab_id: str, safety_type: str, type_name: str, limits: Dict[str, float], direction: str, config: IOConfig):
        self.id = cab_id
        self.safety_type = safety_type
        self.type_name = type_name
        self.limits = limits
        self.direction = direction
        self.config = config
        self.rail_used = 0
        self.channels_relay_di = 0
        self.channels_relay_do = 0
        self.channels_relay_ai = 0
        self.channels_relay_ao = 0
        self.granular_counts = {
            "Barrier_AI": 0,
            "Barrier_AO": 0,
            "Barrier_DI": 0,
            "Barrier_DO": 0,
            "Terminal_AI": 0,
            "Terminal_AO": 0,
            "Terminal_DI": 0,
            "Terminal_DO": 0,
        }
        self.assigned_jbs: List[JBDemand] = []
        self.limiting_factor = "Rail (0.0%)"

    def _required_boards(self, total_channels: int, board_key: str, default_ch_per_board: int):
        ch_per_board = self.config.channels.get(board_key, default_ch_per_board)
        if ch_per_board <= 0:
            return 1 if total_channels > 0 else 0
        return math.ceil(total_channels / ch_per_board) if total_channels > 0 else 0

    def _pool_mode(self):
        max_total = self.limits.get("Max_Total_Boards", None)
        if max_total is None:
            return None, None
        max_total = float(max_total)
        if abs(max_total - 100.0) < 1e-9:
            return "PERCENT_PER_BOARD", 100.0
        return "COUNT", max_total

    def _pool_usage_after(self, jb: JBDemand):
        mode, max_total = self._pool_mode()
        if mode is None:
            return None, None, None

        prefix = "Barrier" if self.safety_type == "IS" else "Terminal"
        usage = 0.0

        for io_suffix in ["AI", "AO", "DI", "DO"]:
            board_key = f"{prefix} Board ({io_suffix})"
            if board_key not in self.limits:
                continue
            granular_key = f"{prefix}_{io_suffix}"
            current_channels = self.granular_counts.get(granular_key, 0)
            added_channels = jb.channel_counts.get(granular_key, 0)
            total_channels = current_channels + added_channels
            req = self._required_boards(total_channels, board_key, 16)
            if mode == "PERCENT_PER_BOARD":
                usage += req * float(self.limits[board_key])
            else:
                usage += req

        relay_map = {
            "AI": ("channels_relay_ai", "Relay Board AI capacity"),
            "AO": ("channels_relay_ao", "Relay Board AO capacity"),
            "DI": ("channels_relay_di", "Relay Board DI capacity"),
            "DO": ("channels_relay_do", "Relay Board DO capacity"),
        }

        for r_type, (attr_name, limit_key) in relay_map.items():
            if limit_key not in self.limits:
                continue
            current = getattr(self, attr_name, 0)
            added = jb.channel_counts.get(f"Relay_{r_type}", 0)
            total = current + added
            req = self._required_boards(total, limit_key, 32)
            if mode == "PERCENT_PER_BOARD":
                usage += req * float(self.limits[limit_key])
            else:
                usage += req

        return usage, max_total, mode

    def can_fit(self, jb: JBDemand):
        if self.safety_type != jb.safety_type:
            return False, "Safety_Mismatch"

        rail_limit = self.limits.get("Max rail terminals", 0)
        if rail_limit > 0 and (self.rail_used + jb.rail_load > rail_limit):
            return False, "Rail_Limit"

        usage, max_total, mode = self._pool_usage_after(jb)

        if mode == "COUNT":
            prefix = "Barrier" if self.safety_type == "IS" else "Terminal"
            for io_suffix in ["AI", "AO", "DI", "DO"]:
                board_key = f"{prefix} Board ({io_suffix})"
                if board_key not in self.limits:
                    continue
                max_boards = int(self.limits[board_key])
                granular_key = f"{prefix}_{io_suffix}"
                current_channels = self.granular_counts.get(granular_key, 0)
                added_channels = jb.channel_counts.get(granular_key, 0)
                total_channels = current_channels + added_channels
                req = self._required_boards(total_channels, board_key, 16)
                if req > max_boards:
                    return False, f"{board_key}_Limit"

            relay_map = {
                "AI": ("channels_relay_ai", "Relay Board AI capacity"),
                "AO": ("channels_relay_ao", "Relay Board AO capacity"),
                "DI": ("channels_relay_di", "Relay Board DI capacity"),
                "DO": ("channels_relay_do", "Relay Board DO capacity"),
            }
            for r_type, (attr_name, limit_key) in relay_map.items():
                if limit_key not in self.limits:
                    continue
                max_boards = int(self.limits[limit_key])
                current = getattr(self, attr_name, 0)
                added = jb.channel_counts.get(f"Relay_{r_type}", 0)
                total = current + added
                req = self._required_boards(total, limit_key, 32)
                if req > max_boards:
                    return False, f"{limit_key}_Limit"

        if mode is not None and max_total is not None:
            if usage > float(max_total) + 1e-9:
                return False, f"{mode}_POOL_FULL (usage={usage:.2f}/{max_total})"

        return True, None

    def add_jb(self, jb: JBDemand):
        self.assigned_jbs.append(jb)
        self.rail_used += jb.rail_load
        self.channels_relay_di += jb.channel_counts.get("Relay_DI", 0)
        self.channels_relay_do += jb.channel_counts.get("Relay_DO", 0)
        self.channels_relay_ai += jb.channel_counts.get("Relay_AI", 0)
        self.channels_relay_ao += jb.channel_counts.get("Relay_AO", 0)
        for key in self.granular_counts.keys():
            self.granular_counts[key] += jb.channel_counts.get(key, 0)


def assign_cabinets(jb_demands: Dict[str, JBDemand], config: IOConfig):
    sorted_jbs = sorted(
        jb_demands.values(),
        key=lambda x: x.rail_load
        + x.channel_counts["Barrier_AI"]
        + x.channel_counts["Barrier_AO"]
        + x.channel_counts["Barrier_DI"]
        + x.channel_counts["Barrier_DO"]
        + x.channel_counts["Terminal_AI"]
        + x.channel_counts["Terminal_AO"]
        + x.channel_counts["Terminal_DI"]
        + x.channel_counts["Terminal_DO"],
        reverse=True,
    )

    cabinets: List[Cabinet] = []
    cabinet_type_pool = build_cabinet_pool_from_plan(config)
    counters = {"IS": 1, "NIS": 1}

    for jb in sorted_jbs:
        assigned = False
        for cab in cabinets:
            if cab.safety_type != jb.safety_type:
                continue
            fits, _ = cab.can_fit(jb)
            if fits:
                cab.add_jb(jb)
                assigned = True
                break
        if assigned:
            continue

        if cabinet_type_pool:
            chosen_idx = -1
            chosen_type = None
            for i, tdef in enumerate(cabinet_type_pool):
                st = jb.safety_type
                temp_cab = Cabinet(
                    cab_id=f"TEMP-{st}-{i}",
                    safety_type=st,
                    type_name=tdef["name"],
                    limits=tdef["limits"],
                    direction=tdef.get("direction", ""),
                    config=config,
                )
                fits, _ = temp_cab.can_fit(jb)
                if fits:
                    chosen_idx = i
                    chosen_type = tdef
                    break
            if chosen_type is None:
                raise RuntimeError(
                    f"JB {jb.name} cannot fit in any remaining defined Cabinet Types."
                )
            cabinet_type_pool.pop(chosen_idx)

            st = jb.safety_type
            direction = chosen_type.get("direction", "")
            display_direction = direction if config.has_directions else ""
            suffix = ""
            if config.has_directions:
                suffix = "R" if direction == "REAR" else ("F" if direction == "FRONT" else "")
            name = f"{config.is_cab_prefix if st == 'IS' else config.nis_cab_prefix}{counters[st]:02d}{suffix}"
            counters[st] += 1

            new_cab = Cabinet(
                cab_id=name,
                safety_type=st,
                type_name=chosen_type["name"],
                limits=chosen_type["limits"],
                direction=display_direction,
                config=config,
            )
            new_cab.add_jb(jb)
            cabinets.append(new_cab)
        else:
            st = jb.safety_type
            name = f"{config.is_cab_prefix if st == 'IS' else config.nis_cab_prefix}{counters[st]:02d}"
            counters[st] += 1

            default_limits = {
                "Max rail terminals": 1000,
                "Max_Total_Boards": 999,
                "Barrier Board (AI)": 999,
                "Barrier Board (AO)": 999,
                "Barrier Board (DI)": 999,
                "Barrier Board (DO)": 999,
                "Terminal Board (AI)": 999,
                "Terminal Board (AO)": 999,
                "Terminal Board (DI)": 999,
                "Terminal Board (DO)": 999,
                "Relay Board AO capacity": 999,
                "Relay Board AI capacity": 999,
                "Relay Board DO capacity": 999,
                "Relay Board DI capacity": 999,
            }
            new_cab = Cabinet(name, st, "Default", default_limits, "", config)
            fits, reason = new_cab.can_fit(jb)
            if not fits:
                raise RuntimeError(f"JB {jb.name} too big even for default cabinet: {reason}")
            new_cab.add_jb(jb)
            cabinets.append(new_cab)

    for cab in cabinets:
        rail_limit = cab.limits.get("Max rail terminals", 0)
        rail_pct = cab.rail_used / rail_limit if rail_limit > 0 else 0.0
        cab.limiting_factor = f"Rail ({rail_pct:.1%})"

        dummy = JBDemand("DUMMY", cab.safety_type)
        dummy.channel_counts = {k: 0 for k in dummy.channel_counts}
        dummy.rail_load = 0

        usage_now, max_total2, mode = cab._pool_usage_after(dummy)
        if usage_now is not None and max_total2 and max_total2 > 0:
            pct = usage_now / float(max_total2)
            if pct > rail_pct:
                cab.limiting_factor = f"{mode} Pool ({pct:.1%})"

    return cabinets


def distribute_evenly(signals: List[Dict[str, Any]], channels_per_board: int):
    total = len(signals)
    if total == 0:
        return [], 0
    num_boards = math.ceil(total / channels_per_board)
    base_count = total // num_boards
    remainder = total % num_boards
    distribution = []
    start_idx = 0
    for b_idx in range(num_boards):
        count = base_count + (1 if b_idx < remainder else 0)
        chunk = signals[start_idx: start_idx + count]
        for sig in chunk:
            distribution.append((b_idx, sig))
        start_idx += count
    return distribution, num_boards


def generate_optimized_output(cabinets: List[Cabinet], config: IOConfig) -> pd.DataFrame:
    final_rows = []
    cabinets.sort(key=lambda c: c.id)
    io_col = config.col_mapping["IO_TYPE"]
    loc_col = config.col_mapping["LOCATION"]

    for cab in cabinets:
        board_seq = {}
        all_signals = []
        for jb in cab.assigned_jbs:
            all_signals.extend(jb.signals)

        df_cab = pd.DataFrame(all_signals)
        if df_cab.empty:
            continue

        def get_board_props(row):
            board_type = "Barrier Board" if cab.safety_type == "IS" else "Terminal Board"
            io = row.get(io_col, "")
            loc = row.get(loc_col, "")
            base_io = io
            if io == "DIR":
                base_io = "DI"
            elif io == "DOR":
                base_io = "DO"
            elif io == "AIR":
                base_io = "AI"
            elif io == "AOR":
                base_io = "AO"
            relay_req = "YES" if (loc == "MCC" and base_io in ["DI", "DO", "AI", "AO"]) else "NO"
            return board_type, relay_req, base_io

        props = df_cab.apply(lambda r: get_board_props(r), axis=1, result_type="expand")
        df_cab["Board_Type"] = props[0]
        df_cab["Relay_Req"] = props[1]
        df_cab["Base_IO"] = props[2]

        df_cab = master_sort(df_cab, config)
        df_cab["Group_Key"] = list(zip(df_cab["Base_IO"], df_cab["Board_Type"]))
        unique_keys = df_cab["Group_Key"].unique()

        for key in unique_keys:
            base_io_type, board_type = key
            group_df = df_cab[df_cab["Group_Key"] == key]
            signals_list = group_df.to_dict("records")
            relay_req = group_df.iloc[0]["Relay_Req"]
            chan_limit_key = f"{board_type} ({base_io_type})"
            chan_limit = config.channels.get(chan_limit_key, 16)

            distributed_signals, num_boards_used = distribute_evenly(signals_list, chan_limit)
            board_groups = {}
            for rel_b_idx, sig in distributed_signals:
                board_groups.setdefault(rel_b_idx, []).append(sig)

            for rel_b_idx in range(num_boards_used):
                chunk = board_groups.get(rel_b_idx, [])
                seq_key = (board_type, base_io_type)
                board_seq[seq_key] = board_seq.get(seq_key, 0) + 1
                actual_board_id = board_seq[seq_key]

                for i, sig in enumerate(chunk):
                    out_row = sig.copy()
                    for dropk in ["Board_Type", "Relay_Req", "Group_Key", "Base_IO"]:
                        out_row.pop(dropk, None)

                    out_row["Cabinet_ID"] = cab.id
                    out_row["Cabinet_Type"] = cab.type_name
                    out_row["Cabinet_Direction"] = cab.direction
                    out_row["Limiting_Factor"] = cab.limiting_factor
                    out_row["Board_ID"] = f"{board_type} ({base_io_type})-{actual_board_id:02d}"
                    out_row["Board_Type"] = board_type
                    out_row["Channel_No"] = i + 1
                    out_row["Relay_Required"] = relay_req
                    final_rows.append(out_row)

                slots_used = len(chunk)
                if slots_used < chan_limit:
                    rep_sig = chunk[-1] if chunk else {
                        config.col_mapping["JB"]: "-",
                        config.col_mapping["IO_TYPE"]: base_io_type,
                        config.col_mapping["SAFETY"]: cab.safety_type,
                        config.col_mapping["LOCATION"]: "-",
                        "Signal_Type": "SPARE",
                    }

                    for ch in range(slots_used + 1, chan_limit + 1):
                        spare_row = rep_sig.copy()
                        for dropk in ["Board_Type", "Relay_Req", "Group_Key", "Base_IO"]:
                            spare_row.pop(dropk, None)

                        spare_row["Tag No"] = "SPARE"
                        spare_row["Loop No"] = "SPARE"
                        spare_row["Signal_Type"] = "SPARE"
                        spare_row[config.col_mapping["TERM1"]] = ""
                        spare_row[config.col_mapping["TERM2"]] = ""
                        spare_row[config.col_mapping["SRC"]] = ""
                        spare_row["Cabinet_ID"] = cab.id
                        spare_row["Cabinet_Type"] = cab.type_name
                        spare_row["Cabinet_Direction"] = cab.direction
                        spare_row["Limiting_Factor"] = cab.limiting_factor
                        spare_row["Board_ID"] = f"{board_type} ({base_io_type})-{actual_board_id:02d}"
                        spare_row["Board_Type"] = board_type
                        spare_row["Channel_No"] = ch
                        spare_row["Relay_Required"] = "NO"
                        final_rows.append(spare_row)

    df_final = pd.DataFrame(final_rows)

    cols = [
        "Cabinet_ID",
        "Cabinet_Type",
        "Cabinet_Direction",
        "Limiting_Factor",
        "Board_ID",
        "Board_Type",
        "Channel_No",
        "Relay_Required",
        "Signal_Type",
        config.col_mapping["JB"],
        config.col_mapping["IO_TYPE"],
        config.col_mapping["TERM1"],
        config.col_mapping["TERM2"],
        "Tag No",
        "Location",
        config.col_mapping["SAFETY"],
    ]
    exists = [c for c in cols if c in df_final.columns]
    remain = [c for c in df_final.columns if c not in exists and c not in ["Rank_IO", "Rank_Safety", "Rank_Loc"]]
    return df_final[exists + remain]


def create_final_excel(final_df: pd.DataFrame, cleaning_report: Dict[str, Any], cabinets: List[Cabinet], output_filename: str, config: IOConfig):
    wb = Workbook()
    fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_orange = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
    bold_font = Font(bold=True)

    ws1 = wb.active
    ws1.title = "IO Assignment"
    for row in dataframe_to_rows(final_df, index=False, header=True):
        ws1.append(row)
    for cell in ws1[1]:
        cell.font = bold_font

    ws2 = wb.create_sheet("input_report")
    ws2.append(["Cleaning Summary"])
    ws2["A1"].font = bold_font
    row_idx = 2
    summary = cleaning_report.get("Summary", {})
    for k, v in summary.items():
        ws2.cell(row=row_idx, column=1, value=k)
        ws2.cell(row=row_idx, column=2, value=v)
        row_idx += 1

    if row_idx > 2:
        pie = PieChart()
        pie.title = "Cleaning Impact"
        labels = Reference(ws2, min_col=1, min_row=2, max_row=row_idx - 1)
        data = Reference(ws2, min_col=2, min_row=2, max_row=row_idx - 1)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        ws2.add_chart(pie, "D2")

    row_idx += 10

    def dump_report_section(title, data_list, color_fill):
        nonlocal row_idx
        ws2.cell(row=row_idx, column=1, value=title).font = bold_font
        row_idx += 1
        if not data_list:
            ws2.cell(row=row_idx, column=1, value="(None)")
            row_idx += 2
            return
        df_sec = pd.DataFrame(data_list)
        headers = df_sec.columns.tolist()
        for col_num, h in enumerate(headers, 1):
            c = ws2.cell(row=row_idx, column=col_num, value=h)
            c.font = bold_font
            if color_fill:
                c.fill = color_fill
        row_idx += 1
        for _, row_data in df_sec.iterrows():
            for col_num, val in enumerate(row_data, 1):
                c = ws2.cell(row=row_idx, column=col_num, value=str(val))
                if color_fill:
                    c.fill = color_fill
            row_idx += 1
        row_idx += 2

    dump_report_section("Removed / Critical (Missing Data)", cleaning_report.get("Removed_Missing_Critical", []), fill_red)
    dump_report_section("Removed Invalid JBs", cleaning_report.get("Removed_Invalid_JB", []), fill_red)
    dump_report_section("Invalid IO Types", cleaning_report.get("Invalid_IO", []), fill_red)
    dump_report_section("Safety Conversions (Warning)", cleaning_report.get("Safety_Conversion", []), fill_orange)
    dump_report_section("Location Imputed", cleaning_report.get("Location_Imputed", []), fill_orange)
    dump_report_section("Logical Exceptions", cleaning_report.get("Logical_Exceptions", []), fill_orange)

    ws3 = wb.create_sheet("output_report")
    ws3.column_dimensions["A"].width = 30

    total_active = len(final_df[final_df["Signal_Type"] == "ACTIVE"])
    total_hot = len(final_df[final_df["Signal_Type"] == "HOT_SPARE"])
    total_spare = len(final_df[final_df["Signal_Type"] == "SPARE"])

    ws3.append(["Global Metrics"])
    ws3["A1"].font = bold_font
    ws3.append(["Total Active Signals", total_active])
    ws3.append(["Actual Hot Spares", total_hot])
    ws3.append(["Expected Hot Spares (approx 20%)", math.ceil(total_active * config.hot_spare_ratio)])
    ws3.append(["Total Spare (Backfill) Channels", total_spare])
    ws3.append([])

    ws3.append(["Per Cabinet Statistics"])
    ws3["A7"].font = bold_font

    headers = [
        "Cabinet", "Type", "Direction", "Limiting Factor",
        "Rail Used (Max)",
        "Barrier AI (Boards)", "Barrier AO (Boards)", "Barrier DI (Boards)", "Barrier DO (Boards)",
        "Terminal AI (Boards)", "Terminal AO (Boards)", "Terminal DI (Boards)", "Terminal DO (Boards)",
        "Relay AI (Boards)", "Relay AO (Boards)", "Relay DI (Boards)", "Relay DO (Boards)",
        "JB Count", "Hot Spares", "Backfill Spares",
    ]
    ws3.append(headers)

    for cab in cabinets:
        hot_cnt = len(final_df[(final_df["Cabinet_ID"] == cab.id) & (final_df["Signal_Type"] == "HOT_SPARE")])
        backfill_cnt = len(final_df[(final_df["Cabinet_ID"] == cab.id) & (final_df["Signal_Type"] == "SPARE")])
        row_data = [
            cab.id,
            cab.type_name,
            cab.direction,
            cab.limiting_factor,
            f"{cab.rail_used} ({cab.limits.get('Max rail terminals', 'N/A')})",
            math.ceil(cab.granular_counts.get("Barrier_AI", 0) / config.channels.get("Barrier Board (AI)", 1)),
            math.ceil(cab.granular_counts.get("Barrier_AO", 0) / config.channels.get("Barrier Board (AO)", 1)),
            math.ceil(cab.granular_counts.get("Barrier_DI", 0) / config.channels.get("Barrier Board (DI)", 1)),
            math.ceil(cab.granular_counts.get("Barrier_DO", 0) / config.channels.get("Barrier Board (DO)", 1)),
            math.ceil(cab.granular_counts.get("Terminal_AI", 0) / config.channels.get("Terminal Board (AI)", 1)),
            math.ceil(cab.granular_counts.get("Terminal_AO", 0) / config.channels.get("Terminal Board (AO)", 1)),
            math.ceil(cab.granular_counts.get("Terminal_DI", 0) / config.channels.get("Terminal Board (DI)", 1)),
            math.ceil(cab.granular_counts.get("Terminal_DO", 0) / config.channels.get("Terminal Board (DO)", 1)),
            math.ceil(cab.channels_relay_ai / config.channels.get("Relay Board AI capacity", 1)),
            math.ceil(cab.channels_relay_ao / config.channels.get("Relay Board AO capacity", 1)),
            math.ceil(cab.channels_relay_di / config.channels.get("Relay Board DI capacity", 1)),
            math.ceil(cab.channels_relay_do / config.channels.get("Relay Board DO capacity", 1)),
            len(cab.assigned_jbs),
            hot_cnt,
            backfill_cnt,
        ]
        ws3.append(row_data)

    wb.save(output_filename)


def run_io_assignment(input_excel_path: str, output_excel_path: str, config_overrides: Optional[Dict[str, Any]] = None):
    logger.info("IO Assignment started: input=%s output=%s", input_excel_path, output_excel_path)
    config = build_config(config_overrides)
    logger.info("Config loaded: plan=%s limits=%s", len(config.cabinet_plan), len(config.cabinet_limits))
    df_raw = pd.read_excel(input_excel_path)
    logger.info("Input rows: %s", len(df_raw))
    df, cleaning_report = initial_cleaning_and_report(df_raw, config)
    logger.info("After cleaning: %s rows", len(df))
    df = master_sort(df, config)
    df = inject_hot_spares(df, config)
    df = master_sort(df, config)
    logger.info("After hot spares: %s rows", len(df))
    demands = calculate_demands(df, config)
    logger.info("JB demands: %s", len(demands))
    cabinets = assign_cabinets(demands, config)
    logger.info("Cabinets assigned: %s", len(cabinets))
    final_df = generate_optimized_output(cabinets, config)
    logger.info("Final output rows: %s", len(final_df))
    create_final_excel(final_df, cleaning_report, cabinets, output_excel_path, config)
    logger.info("Excel saved: %s", output_excel_path)
    return {
        "final_df": final_df,
        "cleaning_report": cleaning_report,
        "cabinets": cabinets,
        "config": config,
    }
